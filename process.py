import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

output_path = './pcr_paperwork.xlsx'


def get_codes_sheet_json():
    input_file = openpyxl.load_workbook('input.xlsx')

    codes_sheet = input_file['Codes']

    max_col = codes_sheet.max_column
    max_row = codes_sheet.max_row

    full_data = []

    code_keys = []
    assay_keys = []

    for col in codes_sheet.iter_cols(min_row=1, min_col=2, max_col=max_col, max_row=1):
        for cell in col:
            code_keys.append(cell.value)

    for row in codes_sheet.iter_rows(min_row=2, min_col=1, max_col=1, max_row=max_row):
        for cell in row:
            assay_keys.append(cell.value)

    code_index = 0
    for col in codes_sheet.iter_cols(min_row=2, min_col=2, max_col=max_col, max_row=max_row):
        data_dict = {'code': code_keys[code_index]}
        code_index += 1

        assay_index = 0
        for cell in col:
            data_dict[assay_keys[assay_index]] = cell.value
            assay_index += 1

        full_data.append(data_dict)

    return full_data


def get_samples_list():
    # read input.xlsx
    input_file = openpyxl.load_workbook('input.xlsx')

    samples = input_file['Samples']

    ext_array = []
    sample_array = []
    assay_array = []

    samples_list = []

    for row in range(2, samples.max_row + 1):
        ext_array.append(samples.cell(row=row, column=1).value)

    for row in range(2, samples.max_row + 1):
        sample_array.append(samples.cell(row=row, column=2).value)

    for row in range(2, samples.max_row + 1):
        assay_array.append(samples.cell(row=row, column=3).value)

    for i in range(len(ext_array)):
        dict_obj = {'ext': ext_array[i], 'id': sample_array[i], 'code': assay_array[i]}
        samples_list.append(dict_obj)

    return samples_list


def generate_pcr_paperwork():

    plate_number = 1

    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    sheet.title = "Plate " + str(plate_number)

    for col in range(3, 15):
        sheet.column_dimensions[get_column_letter(col)].width = 13

        sheet.cell(row=2, column=col).value = col - 2

    plate_letters = ["A", "B", "C", "D", "E", "F", "G", "H"]

    for row in range(3, 11):
        sheet.row_dimensions[row].height = 50
        sheet.cell(row=row, column=2).value = plate_letters[row-3]



    wb.save(output_path)


generate_pcr_paperwork()


