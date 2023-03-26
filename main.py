import json

from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

import openpyxl
import os.path

# check if input file exists in program, if not create it
input_path = './input.xlsx'
input_xlsx = os.path.isfile(input_path)


# create input.xlsx as interface
def create_input_file():

    if not input_xlsx:
        wb = openpyxl.Workbook()
        sheet = wb['Sheet']
        sheet.title = 'Samples'
        wb.create_sheet('Codes')

        # create default samples sheet
        cell_value = PatternFill(start_color='00339966', end_color='00339966', fill_type='solid')

        samples_sheet = wb['Samples']
        codes_sheet = wb['Codes']

        # create headers for samples sheet
        samples_sheet.column_dimensions['A'].width = 20
        samples_sheet.column_dimensions['B'].width = 20
        samples_sheet.column_dimensions['C'].width = 20
        samples_sheet['A1'].fill = PatternFill(fgColor='00C0C0C0', fill_type='solid')
        samples_sheet['B1'].fill = PatternFill(fgColor='00C0C0C0', fill_type='solid')
        samples_sheet['C1'].fill = PatternFill(fgColor='00C0C0C0', fill_type='solid')
        samples_sheet['A1'] = "Extraction Group"
        samples_sheet['B1'] = "Sample ID"
        samples_sheet['C1'] = "Assay Code"
        samples_sheet['A2'] = "Paste Data Here"

        samples_sheet.conditional_formatting.add('A2:C5000',
                                                 FormulaRule(formula=['NOT(ISBLANK(A2))'],
                                                             stopIfTrue=True,
                                                             fill=cell_value))

        # create default codes sheet
        assay_name = PatternFill(start_color='00C0C0C0', end_color='00C0C0C0', fill_type='solid')
        assay_code = PatternFill(start_color='00C0C0C0', end_color='00C0C0C0', fill_type='solid')

        # create headers for codes sheet
        codes_sheet['A1'] = "Assay"
        codes_sheet.column_dimensions['A'].width = 20
        codes_sheet['A1'].fill = PatternFill(fgColor='00808080', fill_type='solid')

        codes_sheet['B1'] = "code1"
        codes_sheet['C1'] = "code2"
        codes_sheet['D1'] = "code3"

        codes_sheet['A2'] = "assay1"
        codes_sheet['A3'] = "assay2"
        codes_sheet['A4'] = "assay3"

        codes_sheet.conditional_formatting.add('B1:ZZ1',
                                               FormulaRule(formula=['NOT(ISBLANK(B1))'],
                                                           stopIfTrue=True,
                                                           fill=assay_code))

        codes_sheet.conditional_formatting.add('A2:A5000',
                                               FormulaRule(formula=['NOT(ISBLANK(A2))'],
                                                           stopIfTrue=True,
                                                           fill=assay_name))

        last_cell_row = codes_sheet.max_row
        last_cell_col = codes_sheet.max_column

        for row in range(2, last_cell_row + 1):
            for col in range(2, last_cell_col + 1):
                codes_sheet.cell(row=row, column=col).value = "N"

        last_cell = codes_sheet.cell(last_cell_row, last_cell_col).coordinate
        editing_area = "B2:" + last_cell

        no = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')
        yes = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')

        codes_sheet.conditional_formatting.add(editing_area,
                                               FormulaRule(formula=['IF(B2="N", "True", "")'],
                                                           stopIfTrue=True,
                                                           fill=no))

        codes_sheet.conditional_formatting.add(editing_area,
                                               FormulaRule(formula=['IF(B2="Y", "True", "")'],
                                                           stopIfTrue=True,
                                                           fill=yes))

        dv = DataValidation(type='list', formula1='"Y,N"', allow_blank=False, showDropDown=False, showErrorMessage=True)
        dv.error = "Invalid Entry - Use Dropdown"

        dv.add(editing_area)

        codes_sheet.add_data_validation(dv)

        wb.save(input_path)


create_input_file()


def get_codes_sheet_json():
    input_file = openpyxl.load_workbook('input.xlsx')

    codes_sheet = input_file['Codes']

    max_col = codes_sheet.max_column
    max_row = codes_sheet.max_row

    full_data = []

    code_keys = []
    assay_keys = []

    for col in codes_sheet.iter_cols(min_row=1, min_col=2, max_col=max_col, max_row=1):
        for cell in col:
            code_keys.append(cell.value)

    for row in codes_sheet.iter_rows(min_row=2, min_col=1, max_col=1, max_row=max_row):
        for cell in row:
            assay_keys.append(cell.value)

    code_index = 0
    for col in codes_sheet.iter_cols(min_row=2, min_col=2, max_col=max_col, max_row=max_row):
        data_dict = {'code': code_keys[code_index]}
        code_index += 1

        assay_index = 0
        for cell in col:
            data_dict[assay_keys[assay_index]] = cell.value
            assay_index += 1

        full_data.append(data_dict)

    print(full_data)

    file_name = 'assay_data.json'
    with open(file_name, 'w') as empty_json:
        json.dump(full_data, empty_json, indent=4)
        empty_json.close()


get_codes_sheet_json()


def update_codes_sheet():
    input_file = openpyxl.load_workbook('input.xlsx')

    codes_sheet = input_file['Codes']

    last_cell_row = codes_sheet.max_row
    last_cell_col = codes_sheet.max_column

    last_cell = codes_sheet.cell(last_cell_row, last_cell_col).coordinate
    editing_area = "B2:" + last_cell

    no = PatternFill(start_color='00FF0000', end_color='00FF0000', fill_type='solid')
    yes = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type='solid')

    codes_sheet.conditional_formatting.add(editing_area,
                                           FormulaRule(formula=['IF(B2="N", "True", "")'],
                                                       stopIfTrue=True,
                                                       fill=no))

    codes_sheet.conditional_formatting.add(editing_area,
                                           FormulaRule(formula=['IF(B2="Y", "True", "")'],
                                                       stopIfTrue=True,
                                                       fill=yes))

    dv = DataValidation(type='list', formula1='"Y,N"', allow_blank=False, showDropDown=False, showErrorMessage=True)
    dv.error = "Invalid Entry - Use Dropdown"

    dv.add(editing_area)

    codes_sheet.add_data_validation(dv)

    input_file.save(input_path)


def get_samples_list():
    # read input.xlsx
    input_file = openpyxl.load_workbook('input.xlsx')

    samples = input_file['Samples']

    ext_array = []
    sample_array = []
    assay_array = []

    samples_list = []

    for row in range(2, samples.max_row + 1):
        ext_array.append(samples.cell(row=row, column=1).value)

    for row in range(2, samples.max_row + 1):
        sample_array.append(samples.cell(row=row, column=2).value)

    for row in range(2, samples.max_row + 1):
        assay_array.append(samples.cell(row=row, column=3).value)

    for i in range(len(ext_array)):
        dict_obj = {'ext': ext_array[i], 'id': sample_array[i], 'code': assay_array[i]}
        samples_list.append(dict_obj)

    return samples_list
